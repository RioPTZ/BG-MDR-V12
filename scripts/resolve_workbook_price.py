from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


def normalize_code(s: str) -> str:
    return re.sub(r'\s+', '', (s or '').upper())


def split_code(code: str):
    m = re.match(r'^([A-Z]+)(\d+)$', normalize_code(code))
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def code_matches_cell(code: str, text: str) -> bool:
    c_norm = normalize_code(code)
    t_norm = normalize_code(text)
    if c_norm in t_norm:
        return True

    prefix, num = split_code(code)
    if prefix is None:
        return False

    compact = re.sub(r'\s+', '', (text or '').upper())
    patterns = [
        rf'{re.escape(prefix)}(\d+)-(\d+)',
        rf'{re.escape(prefix)}(\d+)–(\d+)',
        rf'{re.escape(prefix)}(\d+)—(\d+)',
    ]
    for pat in patterns:
        m = re.search(pat, compact)
        if m:
            start = int(m.group(1))
            end = int(m.group(2))
            if start <= num <= end:
                return True
    return False


def looks_like_price(v) -> bool:
    return isinstance(v, (int, float)) and v >= 0


def main() -> None:
    if len(sys.argv) != 4:
        print('Usage: python3 resolve_workbook_price.py <workbook.xlsx> <code> <system-hint>')
        raise SystemExit(2)

    workbook_path = Path(sys.argv[1])
    code = sys.argv[2]
    system_hint = (sys.argv[3] or '').strip().lower()

    wb = load_workbook(workbook_path, data_only=False)
    candidates = []

    for ws in wb.worksheets:
        header_rows = []
        for r in range(1, min(8, ws.max_row) + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column) + 1)]
            header_rows.append(["" if v is None else str(v) for v in row_vals])

        system_col = None
        if system_hint:
            for c in range(1, min(20, ws.max_column) + 1):
                probe = ' '.join('' if ws.cell(r, c).value is None else str(ws.cell(r, c).value) for r in range(1, min(6, ws.max_row) + 1)).lower()
                if system_hint in probe:
                    system_col = c
                    break

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = ['' if v is None else str(v) for v in row]
            text = ' | '.join(vals)
            if code_matches_cell(code, text):
                price_candidate = None
                if system_col is not None and system_col <= len(row):
                    v = row[system_col - 1]
                    if looks_like_price(v):
                        price_candidate = v
                if price_candidate is None:
                    numeric_values = [v for v in row if looks_like_price(v)]
                    price_candidate = numeric_values[0] if numeric_values else None
                candidates.append({
                    'sheet': ws.title,
                    'rowIndex': i,
                    'matchedText': text,
                    'systemColumn': system_col,
                    'priceCandidate': price_candidate,
                    'headerPreview': header_rows[:4]
                })

    out = {
        'workbook': str(workbook_path),
        'code': code,
        'systemHint': system_hint,
        'candidateCount': len(candidates),
        'candidates': candidates[:20]
    }
    print(json.dumps(out, ensure_ascii=False, indent=2, default=str))


if __name__ == '__main__':
    main()
